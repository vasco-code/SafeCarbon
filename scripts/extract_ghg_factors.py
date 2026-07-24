#!/usr/bin/env python3
"""
Extrai os fatores oficiais da Ferramenta GHG Protocol v2026.0.1 (FGV) e gera o
SQL de seed das tabelas de fator do SafeCarbon (ghg_fuel_factors,
ghg_grid_factors, ghg_gwp, ghg_generic_factors).

Uso:
    python3 scripts/extract_ghg_factors.py /caminho/ferramenta_ghg_protocol_v2026.0.1.xlsx > seed.sql

Os valores usam as colunas JÁ CONVERTIDAS por unidade da planilha (garante
paridade: emissão = quantidade × fator/1000). Não é código de runtime — roda
uma vez para produzir os INSERTs embutidos na migração.
"""
import sys
import openpyxl

SHEET_FE = "Fatores de Emissão"
SHEET_VAR = "Fatores Variáveis"

BIOFUEL_KEYWORDS = [
    "etanol", "bagaço", "bagaco", "biodiesel", "biogás", "biogas", "biometano",
    "biopropano", "carvão vegetal", "carvao vegetal", "lenha", "resíduos vegetais",
    "residuos vegetais", "fração biomassa", "fracao biomassa", "álcool", "alcool",
]


def sqlstr(v):
    if v is None:
        return "null"
    s = str(v).replace("'", "''")
    return f"'{s}'"


def num(v):
    if v is None or v == "-" or v == "":
        return "null"
    try:
        return repr(float(v))
    except (ValueError, TypeError):
        return "null"


def num0(v):
    n = num(v)
    return "0" if n == "null" else n


def is_biofuel(name):
    low = str(name).lower()
    return any(k in low for k in BIOFUEL_KEYWORDS)


def extract_fuels(ws):
    """Tabela 1 + Tabela 2 (combustíveis, rows 31-93). Colunas convertidas T-AB."""
    rows = []
    for r in range(31, 94):
        ref = ws.cell(row=r, column=2).value  # B
        name = ws.cell(row=r, column=3).value  # C
        if not isinstance(ref, (int, float)) or not name or str(name).strip() in ("-", ""):
            continue
        rows.append({
            "ref_no": int(ref),
            "name_pt": str(name).strip(),
            "name_en": ws.cell(row=r, column=4).value,
            "unit": ws.cell(row=r, column=5).value,
            "pci_gj_t": ws.cell(row=r, column=6).value,       # F
            "density_kg_unit": ws.cell(row=r, column=7).value,  # G
            "source_ref": ws.cell(row=r, column=8).value,     # H
            "co2_kg_tj": ws.cell(row=r, column=9).value,      # I
            "co2_kg_un": ws.cell(row=r, column=20).value,     # T
            "ch4_energy": ws.cell(row=r, column=21).value,    # U
            "ch4_manuf": ws.cell(row=r, column=22).value,     # V
            "ch4_comm": ws.cell(row=r, column=23).value,      # W
            "ch4_resid": ws.cell(row=r, column=24).value,     # X
            "n2o_energy": ws.cell(row=r, column=25).value,    # Y
            "n2o_manuf": ws.cell(row=r, column=26).value,     # Z
            "n2o_comm": ws.cell(row=r, column=27).value,      # AA
            "n2o_resid": ws.cell(row=r, column=28).value,     # AB
            "biofuel": is_biofuel(name),
        })
    return rows


def extract_grid(ws):
    """Seção 2 (tabela sazonal, rows 49+): 'FE do SIN' por ano, col Q = média anual."""
    out = []
    for r in range(49, ws.max_row + 1):
        year = ws.cell(row=r, column=2).value       # B
        param = ws.cell(row=r, column=3).value      # C
        annual = ws.cell(row=r, column=17).value    # Q (Média Anual)
        if isinstance(year, (int, float)) and param and str(param).strip().startswith("FE do SIN"):
            if annual is not None and annual != "#N/A":
                out.append({"year": int(year), "co2_t_mwh": annual})
    return out


def extract_gwp(ws):
    """Seção 6, Tabela 23 (rows 581-614): gases controlados por Kyoto."""
    out = []
    for r in range(581, 615):
        label = ws.cell(row=r, column=3).value  # C
        gwp = ws.cell(row=r, column=5).value    # E
        if not label or gwp is None:
            continue
        s = str(label).strip()
        if "(" in s and ")" in s:
            gas = s[s.index("(") + 1:s.index(")")].strip()
        else:
            gas = s
        try:
            float(gwp)
        except (ValueError, TypeError):
            continue
        out.append({"gas": gas, "gwp": gwp})
    return out


def extract_aviation(ws):
    """Tabela 13 (rows 330-332): fatores aéreos convertidos por p.km (H/I/J)."""
    keys = {330: "air_short", 331: "air_medium", 332: "air_long"}
    descs = {330: "Viagem aérea — curta distância (≤ 500 km)",
             331: "Viagem aérea — média distância (500–3700 km)",
             332: "Viagem aérea — longa distância (> 3700 km)"}
    out = []
    for r, key in keys.items():
        out.append({
            "factor_key": key,
            "description": descs[r],
            "co2_kg": ws.cell(row=r, column=8).value,   # H kg CO2/p.km
            "ch4_kg": ws.cell(row=r, column=9).value,   # I kg CH4/p.km
            "n2o_kg": ws.cell(row=r, column=10).value,  # J kg N2O/p.km
        })
    return out


def extract_wtt_fuels(ws):
    """Tabela 22 (Seção 5, 'berço ao portão'/cradle-to-gate), rows 542+, até
    a próxima Tabela/Seção. Colunas F/G/H já em gCO2/MJ = kg/GJ (sem
    conversão). Pula linhas com fator #N/A (calculado a partir de composição
    em outra aba, fora do escopo desta extração)."""
    out = []
    end = None
    for r in range(540, ws.max_row + 1):
        v = ws.cell(row=r, column=2).value or ws.cell(row=r, column=3).value
        if isinstance(v, str) and (v.strip().lower().startswith("tabela 23") or v.strip().lower().startswith("seção 6")):
            end = r
            break
    for r in range(542, end or ws.max_row):
        name = ws.cell(row=r, column=3).value
        co2 = ws.cell(row=r, column=6).value
        ch4 = ws.cell(row=r, column=7).value
        n2o = ws.cell(row=r, column=8).value
        if not name or not isinstance(name, str):
            continue
        try:
            float(co2)
        except (ValueError, TypeError):
            continue
        out.append({"name_pt": name.strip(), "co2_kg_gj": co2, "ch4_kg_gj": ch4, "n2o_kg_gj": n2o})
    return out


def extract_commuting(ws):
    """Starter set: ônibus municipal (Tabela 9, ano mais recente 2025, diesel).
    Fatores convertidos por p.km (F/G/H). Expandir com as demais tabelas de
    transporte de passageiros (10/15/16/18) em fase seguinte."""
    out = []
    # Tabela 9, linha 2025 = row 270; diesel F/G/H
    out.append({
        "factor_key": "bus_municipal_diesel",
        "description": "Ônibus municipal a diesel (por passageiro.km)",
        "co2_kg": ws.cell(row=270, column=6).value,   # F
        "ch4_kg": ws.cell(row=270, column=7).value,   # G
        "n2o_kg": ws.cell(row=270, column=8).value,   # H
    })
    return out


def print_wtt_seed(fe):
    """Seed isolado de ghg_wtt_fuel_factors (Escopo 3 Cat. 3) — usado em
    migrations incrementais que só adicionam essa tabela, sem re-seedar as
    4 tabelas da Fase 1 (já aplicadas)."""
    wtt = extract_wtt_fuels(fe)
    print("-- Seed de ghg_wtt_fuel_factors (WTT/cradle-to-gate, Escopo 3 Cat. 3) — gerado por scripts/extract_ghg_factors.py --wtt")
    print(f"-- {len(wtt)} combustíveis\n")
    print("delete from ghg_wtt_fuel_factors;\n")
    for w in wtt:
        print(
            "insert into ghg_wtt_fuel_factors (name_pt, co2_kg_gj, ch4_kg_gj, n2o_kg_gj, source) values ("
            f"{sqlstr(w['name_pt'])}, {num0(w['co2_kg_gj'])}, {num0(w['ch4_kg_gj'])}, {num0(w['n2o_kg_gj'])}, "
            "'JEC/Ecoinvent via GHG Protocol FGV v2026.0.1');"
        )


def extract_effluent_factors(listas):
    """Fatores de tratamento de efluentes (Escopo 1) — aba "Listas",
    eflu_tipo_tratamento_MCF_domestico (CA3:CE14) e _industrial (CA17:CE25).
    Colunas: CA=tipo, CB=MCF, CC=EF kgCH4/kgDBO, CD=EF kgCH4/kgDQO, CE=kgN2O-N/kgN."""
    rows = []
    for domain, r1, r2 in (("domestic", 3, 14), ("industrial", 17, 25)):
        for r in range(r1, r2 + 1):
            name = listas.cell(row=r, column=79).value  # CA
            if not name or str(name).strip() in ("-", "", "Tipos de tratamento"):
                continue
            rows.append({
                "domain": domain,
                "treatment_type": str(name).strip(),
                "mcf": listas.cell(row=r, column=80).value,
                "ef_ch4_kg_dbo": listas.cell(row=r, column=81).value,
                "ef_ch4_kg_dqo": listas.cell(row=r, column=82).value,
                "ef_n2o_n_kg_n": listas.cell(row=r, column=83).value,
            })
    return rows


def print_effluent_seed(listas):
    """Seed isolado de ghg_effluent_factors (Escopo 1, aba Efluentes)."""
    efl = extract_effluent_factors(listas)
    print("-- Seed de ghg_effluent_factors (tratamento de efluentes, Escopo 1) — gerado por scripts/extract_ghg_factors.py --effluent")
    print(f"-- {len(efl)} tipos de tratamento (domésticos + industriais)\n")
    print("delete from ghg_effluent_factors;\n")
    for e in efl:
        print(
            "insert into ghg_effluent_factors (domain, treatment_type, mcf, ef_ch4_kg_dbo, ef_ch4_kg_dqo, ef_n2o_n_kg_n, source) values ("
            f"{sqlstr(e['domain'])}, {sqlstr(e['treatment_type'])}, {num0(e['mcf'])}, {num0(e['ef_ch4_kg_dbo'])}, "
            f"{num0(e['ef_ch4_kg_dqo'])}, {num0(e['ef_n2o_n_kg_n'])}, 'IPCC 2006 via GHG Protocol FGV v2026.0.1');"
        )


def main():
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    flags = [a for a in sys.argv[1:] if a.startswith("--")]
    path = args[0] if args else "/Users/vasco/Downloads/ferramenta_ghg_protocol_v2026.0.1.xlsx"
    wb = openpyxl.load_workbook(path, data_only=True)
    fe = wb[SHEET_FE]
    var = wb[SHEET_VAR]

    if "--wtt" in flags:
        print_wtt_seed(fe)
        return

    if "--effluent" in flags:
        print_effluent_seed(wb["Listas"])
        return

    fuels = extract_fuels(fe)
    grid = extract_grid(var)
    gwp = extract_gwp(fe)
    aviation = extract_aviation(fe)
    commuting = extract_commuting(fe)

    print("-- Seed de fatores GHG Protocol v2026.0.1 (FGV) — gerado por scripts/extract_ghg_factors.py")
    print(f"-- {len(fuels)} combustíveis, {len(grid)} anos de FE do SIN, {len(gwp)} GWP, "
          f"{len(aviation)} fatores aéreos, {len(commuting)} fatores casa-trabalho\n")

    print("delete from ghg_fuel_factors; delete from ghg_grid_factors; delete from ghg_gwp; delete from ghg_generic_factors;\n")

    for f in fuels:
        print(
            "insert into ghg_fuel_factors (ref_no, name_pt, name_en, unit, pci_gj_t, density_kg_unit, "
            "co2_kg_tj, co2_kg_un, ch4_kg_un_energy, ch4_kg_un_manufacturing, ch4_kg_un_commercial, "
            "ch4_kg_un_residential, n2o_kg_un_energy, n2o_kg_un_manufacturing, n2o_kg_un_commercial, "
            "n2o_kg_un_residential, is_biofuel, source_ref, source) values ("
            f"{f['ref_no']}, {sqlstr(f['name_pt'])}, {sqlstr(f['name_en'])}, {sqlstr(f['unit'])}, "
            f"{num(f['pci_gj_t'])}, {num(f['density_kg_unit'])}, {num(f['co2_kg_tj'])}, "
            f"{num0(f['co2_kg_un'])}, {num0(f['ch4_energy'])}, {num0(f['ch4_manuf'])}, {num0(f['ch4_comm'])}, "
            f"{num0(f['ch4_resid'])}, {num0(f['n2o_energy'])}, {num0(f['n2o_manuf'])}, {num0(f['n2o_comm'])}, "
            f"{num0(f['n2o_resid'])}, {str(f['biofuel']).lower()}, {sqlstr(f['source_ref'])}, 'GHG Protocol FGV v2026.0.1');"
        )
    print()

    for g in grid:
        print(
            "insert into ghg_grid_factors (year, month, region, method, co2_t_mwh, ch4_t_mwh, n2o_t_mwh, source) values ("
            f"{g['year']}, null, 'SIN', 'location', {num0(g['co2_t_mwh'])}, 0, 0, 'MCTI/SIN via GHG Protocol FGV');"
        )
    print()

    for g in gwp:
        print(f"insert into ghg_gwp (gas, gwp, ar_version) values ({sqlstr(g['gas'])}, {num(g['gwp'])}, 'AR5');")
    print()

    for a in aviation:
        print(
            "insert into ghg_generic_factors (source_category, factor_key, description, unit, co2_kg, ch4_kg, n2o_kg, "
            "co2e_kg, biogenic_co2_kg, source) values ("
            f"'business_travel', {sqlstr(a['factor_key'])}, {sqlstr(a['description'])}, 'kg/p.km', "
            f"{num0(a['co2_kg'])}, {num0(a['ch4_kg'])}, {num0(a['n2o_kg'])}, null, 0, 'DEFRA via GHG Protocol FGV');"
        )
    for c in commuting:
        print(
            "insert into ghg_generic_factors (source_category, factor_key, description, unit, co2_kg, ch4_kg, n2o_kg, "
            "co2e_kg, biogenic_co2_kg, source) values ("
            f"'commuting', {sqlstr(c['factor_key'])}, {sqlstr(c['description'])}, 'kg/p.km', "
            f"{num0(c['co2_kg'])}, {num0(c['ch4_kg'])}, {num0(c['n2o_kg'])}, null, 0, 'GHG Protocol FGV v2026.0.1');"
        )


if __name__ == "__main__":
    main()
